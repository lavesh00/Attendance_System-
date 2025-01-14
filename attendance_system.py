import os
import cv2
import face_recognition
import pickle
from datetime import datetime
import openpyxl

# Load encodings
encodings_file = "C:\\Users\\asus\\Downloads\\AI\\attendance_system\\encodings.pkl"
try:
    with open(encodings_file, 'rb') as f:
        known_face_encodings = pickle.load(f)
except FileNotFoundError:
    print("Encodings file not found. Please ensure the file exists.")
    exit()

# Initialize attendance sheet
attendance_file = "C:\\Users\\asus\\Downloads\\AI\\attendance_system\\attendance.xlsx"

if not os.path.exists(attendance_file):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Attendance"
    sheet.append(["Student ID", "Time"])
    workbook.save(attendance_file)

# Load attendance sheet
workbook = openpyxl.load_workbook(attendance_file)
sheet = workbook["Attendance"]

# Create a set of already marked student IDs for quick lookup
marked_students = {row[0].value for row in sheet.iter_rows(min_row=2, max_col=1) if row[0].value}

# Initialize camera
video_capture = cv2.VideoCapture(0)
if not video_capture.isOpened():
    print("Unable to access the camera. Please check your setup.")
    exit()

try:
    while True:
        # Capture frame-by-frame
        ret, frame = video_capture.read()
        if not ret:
            print("Failed to capture frame. Exiting...")
            break

        # Convert the frame to RGB
        rgb_frame = frame[:, :, ::-1]

        # Find all face locations and encodings in the frame
        face_locations = face_recognition.face_locations(rgb_frame)
        face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)

        for face_encoding, face_location in zip(face_encodings, face_locations):
            matches = face_recognition.compare_faces(list(known_face_encodings.values()), face_encoding)
            face_distances = face_recognition.face_distance(list(known_face_encodings.values()), face_encoding)

            if any(matches):
                # Get the best match
                best_match_index = face_distances.argmin()
                student_id = list(known_face_encodings.keys())[best_match_index]

                if student_id not in marked_students:
                    # Record attendance
                    now = datetime.now()
                    time_string = now.strftime("%H:%M:%S")

                    print(f"Marking attendance for {student_id} at {time_string}")
                    sheet.append([student_id, time_string])
                    marked_students.add(student_id)
                    workbook.save(attendance_file)

            # Draw a rectangle around the face
            top, right, bottom, left = face_location
            cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)

        # Display the frame
        cv2.imshow('Attendance System', frame)

        # Break on pressing 'q'
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

finally:
    # Release camera and close OpenCV windows
    video_capture.release()
    cv2.destroyAllWindows()
    print("Resources released. Exiting...")
