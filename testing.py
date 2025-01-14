import os
import openpyxl

# Initialize attendance sheet
attendance_file = "C:\\Users\\asus\\Downloads\\AI\\attendance_system\\attendance.xlsx"

# Check if attendance file exists and is valid
def initialize_attendance_file(file_path):
    try:
        if os.path.exists(file_path):
            # Attempt to load the workbook
            workbook = openpyxl.load_workbook(file_path)
            if "Attendance" not in workbook.sheetnames:
                sheet = workbook.create_sheet("Attendance")
                sheet.append(["Student ID", "Time"])
                workbook.save(file_path)
        else:
            # Create a new workbook
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Attendance"
            sheet.append(["Student ID", "Time"])
            workbook.save(file_path)
        return workbook
    except Exception as e:
        print(f"Error loading or creating Excel file: {e}")
        print("Recreating attendance file...")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Attendance"
        sheet.append(["Student ID", "Time"])
        workbook.save(file_path)
        return workbook

# Initialize the attendance workbook
workbook = initialize_attendance_file(attendance_file)
sheet = workbook["Attendance"]
